# ============================================================
#  ALPHA MODEL — AUTOMATED DATA COLLECTION SCRIPT
#  Pulls: adjusted price, market cap, approx. P/E
#  Output: alpha_data.xlsx  (ready for Stata)
# ============================================================
#
#  HOW TO RUN THIS FILE (complete beginner steps):
#
#  1. Open your terminal / command prompt
#  2. Type:  pip install yfinance openpyxl pandas
#     then press Enter and wait for it to finish
#  3. Save this file somewhere easy, e.g. your Desktop
#  4. In the terminal, navigate to that folder:
#       Mac/Linux:  cd ~/Desktop
#       Windows:    cd C:\Users\YourName\Desktop
#  5. Type:  python collect_alpha_data.py
#     then press Enter
#  6. Wait ~60 seconds — the file alpha_data.xlsx will appear
#     in the same folder as this script
#
# ============================================================

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ── 1. SETTINGS ─────────────────────────────────────────────
START_DATE  = "2020-01-01"
END_DATE    = "2024-12-31"
OUTPUT_FILE = "alpha_data.xlsx"

STOCKS = {
    "AAPL":  "Technology",
    "MSFT":  "Technology",
    "NVDA":  "Technology",
    "AMZN":  "Consumer Disc.",
    "META":  "Comm. Services",
    "GOOGL": "Comm. Services",
    "JPM":   "Financials",
    "GS":    "Financials",
    "V":     "Financials",
    "MA":    "Financials",
    "LULU":  "Consumer Disc.",
    "NKE":   "Consumer Disc.",
    "COST":  "Consumer Staples",
    "WMT":   "Consumer Staples",
    "ADBE":  "Technology",
    "CRM":   "Technology",
    "AMD":   "Technology",
    "AMAT":  "Technology",
    "PYPL":  "Financials",
    "NFLX":  "Comm. Services",
    "TSLA":  "Consumer Disc.",
    "UNH":   "Health Care",
    "LLY":   "Health Care",
    "JNJ":   "Health Care",
    "XOM":   "Energy",
    "CVX":   "Energy",
    "CAT":   "Industrials",
    "HON":   "Industrials",
    "NEE":   "Utilities",
    "AMT":   "Real Estate",
}

# ── 2. DOWNLOAD DATA ─────────────────────────────────────────
print("\n========================================")
print("  ALPHA MODEL DATA COLLECTION SCRIPT")
print("========================================")
print(f"\nDownloading data for {len(STOCKS)} stocks...")
print(f"Period: {START_DATE} to {END_DATE}\n")

all_rows = []
failed   = []

for i, (ticker, sector) in enumerate(STOCKS.items(), 1):
    print(f"  [{i:02d}/{len(STOCKS)}]  {ticker:<6} ...", end=" ")
    try:
        stock = yf.Ticker(ticker)

        # Monthly adjusted close prices
        hist = stock.history(start=START_DATE, end=END_DATE,
                             interval="1mo", auto_adjust=True)
        if hist.empty:
            print("NO DATA — skipped")
            failed.append(ticker)
            continue

        hist.index = hist.index.to_period("M").to_timestamp()
        price_series = hist["Close"].dropna()

        # Approximate P/E: price / trailing EPS
        # EPS is the most recently available trailing 12-month figure.
        # Because price varies each month, P/E still varies month-to-month.
        # EPS only refreshes quarterly — disclose this in your alpha report.
        trailing_eps = stock.info.get("trailingEps", None)
        pe_series    = None
        if trailing_eps and trailing_eps > 0:
            pe_series = (price_series / trailing_eps).clip(lower=0, upper=500)

        # Market cap: shares outstanding x monthly price
        shares = stock.info.get("sharesOutstanding", None)

        for date, price in price_series.items():
            date_str = date.strftime("%Y-%m")
            pe_val   = round(float(pe_series[date]), 2) if (pe_series is not None and date in pe_series.index) else ""
            mktcap   = round((shares * float(price)) / 1e9, 2) if shares else ""

            all_rows.append({
                "date":      date_str,
                "ticker":    ticker,
                "price":     round(float(price), 4),
                "pe_ratio":  pe_val,
                "mktcap_bn": mktcap,
                "sector":    sector,
            })

        print(f"OK  ({len(price_series)} months)")

    except Exception as e:
        print(f"ERROR — {e}")
        failed.append(ticker)

# ── 3. BUILD DATAFRAME ───────────────────────────────────────
df = pd.DataFrame(all_rows)
df = df.sort_values(["ticker", "date"]).reset_index(drop=True)

print(f"\n{'─'*40}")
print(f"  Total rows collected:  {len(df):,}")
print(f"  Stocks with data:      {df['ticker'].nunique()} / {len(STOCKS)}")
if failed:
    print(f"  Failed tickers:        {', '.join(failed)}")
print(f"{'─'*40}\n")

if df.empty:
    print("ERROR: No data collected. Check your internet connection.")
    exit()

# ── 4. WRITE TO EXCEL WITH FORMATTING ───────────────────────
print(f"Writing to {OUTPUT_FILE} ...", end=" ")

DARK_NAVY   = "0D1B2A"
MID_NAVY    = "1B2E45"
ACCENT_BLUE = "2563EB"
LIGHT_BLUE  = "DBEAFE"
PALE_BLUE   = "EFF6FF"
WHITE       = "FFFFFF"
MID_GRAY    = "CBD5E1"
INPUT_BLUE  = "0000FF"

def cfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def tborder():
    s = Side(style="thin", color=MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(c, text, bg=MID_NAVY, sz=9, bold=True):
    c.value = text
    c.font  = Font(name="Arial", size=sz, bold=bold, color=WHITE)
    c.fill  = cfill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = tborder()

wb = Workbook()

# ── Sheet 1: Raw_Data ────────────────────────────────────────
ws = wb.active
ws.title = "Raw_Data"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

for col, w in {"A":12,"B":10,"C":14,"D":13,"E":15,"F":18}.items():
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 36
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = "ALPHA MODEL — RAW DATA  (auto-collected via yfinance)"
c.font  = Font(name="Arial", size=13, bold=True, color=WHITE)
c.fill  = cfill(DARK_NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[2].height = 18
ws.merge_cells("A2:F2")
c = ws["A2"]
c.value = (f"  {df['ticker'].nunique()} stocks  |  "
           f"{df['date'].min()} to {df['date'].max()}  |  "
           f"{len(df):,} observations  |  "
           "P/E = price / trailing EPS  (see Notes_PE_Method tab)")
c.font  = Font(name="Arial", size=9, italic=True, color=WHITE)
c.fill  = cfill(MID_NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")

ws.row_dimensions[3].height = 32
hdrs = [("date","YYYY-MM"),("ticker","Stock code"),("price","Adj. close ($)"),
        ("pe_ratio","P/E (approx.)"),("mktcap_bn","Mkt cap ($bn)"),("sector","GICS sector")]
for ci, (h, sub) in enumerate(hdrs, 1):
    hdr_cell(ws[f"{get_column_letter(ci)}3"], f"{h}\n{sub}")

ticker_seen = {}
for row_idx, row_data in df.iterrows():
    er = row_idx + 4
    ws.row_dimensions[er].height = 14
    tk = row_data["ticker"]
    if tk not in ticker_seen:
        ticker_seen[tk] = len(ticker_seen) % 2 == 0
    bg = PALE_BLUE if ticker_seen[tk] else WHITE

    for ci, val in enumerate([row_data["date"], row_data["ticker"], row_data["price"],
                               row_data["pe_ratio"], row_data["mktcap_bn"], row_data["sector"]], 1):
        c = ws[f"{get_column_letter(ci)}{er}"]
        c.value  = val if val != "" else None
        c.fill   = cfill(bg)
        c.border = tborder()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font   = Font(name="Arial", size=9,
                        color=DARK_NAVY if ci in (1,2,6) else INPUT_BLUE)

# ── Sheet 2: Summary ─────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False
for col, w in {"A":3,"B":14,"C":18,"D":12,"E":14,"F":14,"G":16}.items():
    ws2.column_dimensions[col].width = w

ws2.row_dimensions[1].height = 36
ws2.merge_cells("B1:G1")
c = ws2["B1"]
c.value = "DATA SUMMARY — Coverage check before importing to Stata"
c.font  = Font(name="Arial", size=12, bold=True, color=WHITE)
c.fill  = cfill(DARK_NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

sum_hdrs = ["Ticker","Sector","Months","Avg Price ($)","Avg P/E","Avg MktCap ($bn)"]
ws2.row_dimensions[2].height = 22
for ci, h in enumerate(sum_hdrs, 2):
    hdr_cell(ws2[f"{get_column_letter(ci)}2"], h, bg=ACCENT_BLUE)

def safe_mean(s):
    nums = pd.to_numeric(s, errors="coerce").dropna()
    return round(nums.mean(), 1) if len(nums) > 0 else None

summary = df.groupby("ticker").agg(
    sector    =("sector","first"),
    months    =("date","count"),
    avg_price =("price","mean"),
).reset_index()
summary["avg_pe"]     = [safe_mean(df[df.ticker==t]["pe_ratio"])     for t in summary["ticker"]]
summary["avg_mktcap"] = [safe_mean(df[df.ticker==t]["mktcap_bn"])    for t in summary["ticker"]]

for i, (_, row) in enumerate(summary.iterrows(), 1):
    r  = i + 2
    bg = PALE_BLUE if i % 2 == 0 else WHITE
    ws2.row_dimensions[r].height = 16
    vals = [row["ticker"], row["sector"], int(row["months"]),
            round(row["avg_price"],2), row["avg_pe"], row["avg_mktcap"]]
    for ci, val in enumerate(vals, 2):
        c = ws2[f"{get_column_letter(ci)}{r}"]
        c.value  = val
        c.fill   = cfill(bg)
        c.border = tborder()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font   = Font(name="Arial", size=9,
                        bold=(ci==2), color=ACCENT_BLUE if ci==2 else "000000")

# ── Sheet 3: Notes on P/E method ────────────────────────────
ws3 = wb.create_sheet("Notes_PE_Method")
ws3.sheet_view.showGridLines = False
for col, w in {"A":3,"B":26,"C":56}.items():
    ws3.column_dimensions[col].width = w

ws3.row_dimensions[1].height = 36
ws3.merge_cells("B1:C1")
c = ws3["B1"]
c.value = "NOTES — P/E Approximation Method & Disclosure"
c.font  = Font(name="Arial", size=12, bold=True, color=WHITE)
c.fill  = cfill(DARK_NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")

notes = [
    ("METHOD USED",
     "pe_ratio = monthly_price / trailing_EPS_ttm\n"
     "Trailing EPS is the most recently available 12-month figure from yfinance."),
    ("WHY THIS IS VALID",
     "Price changes every month, so pe_ratio still varies month-by-month.\n"
     "The dominant source of P/E variation is price movement, which is captured.\n"
     "This is a well-known approximation used in quantitative research."),
    ("LIMITATION",
     "EPS only refreshes quarterly, not monthly.\n"
     "P/E between earnings releases will slightly overstate or understate true P/E.\n"
     "This is acceptable for a first model — disclose it clearly."),
    ("DISCLOSURE LANGUAGE FOR YOUR REPORT",
     "\"P/E is approximated as monthly closing price divided by the trailing "
     "twelve-month EPS. EPS refreshes quarterly rather than monthly, which "
     "means the earnings component is held constant between earnings releases. "
     "This is a limitation of the current specification and would be improved "
     "in a production model using quarterly EPS from Compustat or Macrotrends.\""),
    ("STATA: CONVERT TO EARNINGS YIELD",
     "gen ep = 1 / pe_ratio if pe_ratio > 0\n"
     "// Use ep as your valuation signal, not raw pe_ratio\n"
     "// Higher ep = cheaper stock = potentially more attractive\n"
     "// Drop rows where pe_ratio <= 0 (negative earnings = meaningless P/E)"),
    ("HOW TO IMPROVE LATER",
     "1. Download quarterly EPS history from Macrotrends for each ticker\n"
     "2. Merge on quarter, forward-fill within each quarter\n"
     "3. This gives a properly time-varying earnings yield signal"),
]

r = 3
for label, text in notes:
    ws3.row_dimensions[r].height = 20
    b = ws3[f"B{r}"]
    b.value = label
    b.font  = Font(name="Arial", size=9, bold=True, color=WHITE)
    b.fill  = cfill(ACCENT_BLUE)
    b.alignment = Alignment(horizontal="left", vertical="center")
    b.border = tborder()

    lines = text.count("\n") + 1
    ws3.row_dimensions[r].height = max(20, lines * 14)
    c2 = ws3[f"C{r}"]
    c2.value = text
    c2.font  = Font(name="Arial", size=9, color="1E3A5F")
    c2.fill  = cfill(PALE_BLUE)
    c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c2.border = tborder()
    ws3[f"B{r}"].border = tborder()
    r += 1

wb.save(OUTPUT_FILE)
print(f"Done!")
print(f"\n{'='*42}")
print(f"  File saved:  {OUTPUT_FILE}")
print(f"  Open it and check the Summary tab first.")
print(f"  Stata import command:")
print(f'  import excel "{OUTPUT_FILE}", \\')
print(f'    sheet("Raw_Data") firstrow clear')
print(f"{'='*42}\n")
